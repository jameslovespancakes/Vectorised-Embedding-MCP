"""Spreadsheet parser. .xlsx/.xlsm via openpyxl; legacy .xls via xlrd.

One text block emitted per sheet (page=None — pages don't apply to spreadsheets).
Each row joined by ` | `, rows separated by newlines, sheet name prefixed.
Empty cells skipped. Date/number cells stringified via str().
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterator

logger = logging.getLogger(__name__)


def _parse_xlsx(path: Path) -> Iterator[tuple[str, None]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed — skipping %s", path)
        return

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                yield f"[Sheet: {sheet_name}]\n" + "\n".join(rows), None
    finally:
        wb.close()


def _parse_xls(path: Path) -> Iterator[tuple[str, None]]:
    try:
        import xlrd
    except ImportError:
        logger.warning("xlrd not installed — skipping %s", path)
        return

    book = xlrd.open_workbook(str(path))
    for sheet in book.sheets():
        rows: list[str] = []
        for r in range(sheet.nrows):
            cells = []
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if val is None:
                    continue
                s = str(val).strip()
                if s:
                    cells.append(s)
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            yield f"[Sheet: {sheet.name}]\n" + "\n".join(rows), None


def parse(path: Path) -> Iterator[tuple[str, None]]:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        yield from _parse_xlsx(path)
    elif ext == ".xls":
        yield from _parse_xls(path)
    else:
        raise ValueError(f"spreadsheet parser does not handle extension: {ext}")
